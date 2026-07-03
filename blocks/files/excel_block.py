"""
Bloco Excel do PyFlow RPA.
Lê e escreve células em arquivos .xlsx usando openpyxl.
Coloque em: blocks/files/excel_block.py

Requisitos:
  pip install openpyxl
"""
import os
from blocks.base_block import BaseBlock


class ExcelBlock(BaseBlock):
    name        = "Excel (.xlsx)"
    description = "Lê e escreve dados em arquivos Excel (.xlsx) usando openpyxl. Suporta ler célula, ler coluna, ler linha, ler planilha completa, escrever célula e adicionar linha."
    category    = "Arquivos"

    params_schema = [
        {
            "name":    "action",
            "label":   "Ação",
            "type":    "select",
            "required": True,
            "default": "read_cell",
            "options": [
                {"value": "read_cell",   "label": "Ler célula"},
                {"value": "read_column", "label": "Ler coluna inteira"},
                {"value": "read_row",    "label": "Ler linha inteira"},
                {"value": "read_sheet",  "label": "Ler planilha completa"},
                {"value": "write_cell",  "label": "Escrever em célula"},
                {"value": "append_row",  "label": "Adicionar linha no final"},
                {"value": "create",      "label": "Criar arquivo novo"},
            ],
        },
        {
            "name":        "filepath",
            "label":       "Caminho do arquivo .xlsx",
            "type":        "str",
            "required":    True,
            "default":     "",
            "placeholder": "Ex: dados/planilha.xlsx | relatorio.xlsx"
        },
        {
            "name":        "sheet",
            "label":       "Nome da planilha (aba)",
            "type":        "str",
            "required":    False,
            "default":     "",
            "placeholder": "Vazio = primeira aba. Ex: Plan1 | Dados | Sheet1"
        },
        {
            "name":        "cell",
            "label":       "Célula (para read_cell / write_cell)",
            "type":        "str",
            "required":    False,
            "default":     "",
            "placeholder": "Ex: A1 | B3 | C10"
        },
        {
            "name":        "column",
            "label":       "Coluna (para read_column) ou cabeçalho (append_row)",
            "type":        "str",
            "required":    False,
            "default":     "",
            "placeholder": "Ex: A | B | 1 (número) | Nome;Idade;Email (cabeçalhos append)"
        },
        {
            "name":        "row",
            "label":       "Linha (para read_row)",
            "type":        "str",
            "required":    False,
            "default":     "",
            "placeholder": "Ex: 1 | 5 | 10"
        },
        {
            "name":        "value",
            "label":       "Valor a escrever (para write_cell / append_row)",
            "type":        "str",
            "required":    False,
            "default":     "",
            "placeholder": "Ex: {{variavel}} | Texto fixo | val1;val2;val3 (append_row: separar por ;)"
        },
        {
            "name":        "skip_header",
            "label":       "Pular primeira linha ao ler (cabeçalho)",
            "type":        "bool",
            "required":    False,
            "default":     True
        },
        {
            "name":        "variable_name",
            "label":       "Salvar resultado como variável",
            "type":        "str",
            "required":    False,
            "default":     "excel_resultado",
            "placeholder": "Nome da variável para o resultado lido"
        },
    ]

    ACTIONS = {"read_cell", "read_column", "read_row", "read_sheet", "write_cell", "append_row", "create"}

    def execute(self, params: dict) -> dict:
        errors = self.validate_params(params)
        if errors:
            return {"success": False, "message": "\n".join(errors)}

        try:
            import openpyxl
        except ImportError:
            return {"success": False, "message": "openpyxl não instalado. Rode: pip install openpyxl"}

        action      = params.get("action", "read_cell").strip().lower()
        filepath    = params.get("filepath", "").strip()
        sheet_name  = params.get("sheet", "").strip()
        cell        = params.get("cell", "").strip().upper()
        column      = params.get("column", "").strip()
        row         = params.get("row", "").strip()
        value       = params.get("value", "")
        skip_header = params.get("skip_header", True)
        var_name    = params.get("variable_name", "excel_resultado").strip() or "excel_resultado"

        if action not in self.ACTIONS:
            return {"success": False, "message": f"Ação '{action}' inválida. Use: {', '.join(sorted(self.ACTIONS))}"}

        if action == "create":
            return self._create(filepath, openpyxl)

        if not filepath:
            return {"success": False, "message": "filepath é obrigatório."}

        # Para ações de escrita, cria o arquivo se não existir
        if action in {"write_cell", "append_row"} and not os.path.exists(filepath):
            os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
            wb = openpyxl.Workbook()
            wb.save(filepath)

        if not os.path.exists(filepath):
            return {"success": False, "message": f"Arquivo não encontrado: {filepath}"}

        try:
            if action in {"read_cell", "read_column", "read_row", "read_sheet"}:
                return self._read(action, filepath, sheet_name, cell, column,
                                  row, skip_header, var_name, openpyxl)
            else:
                return self._write(action, filepath, sheet_name, cell,
                                   column, value, openpyxl)
        except Exception as e:
            return {"success": False, "message": f"Erro ao processar Excel: {str(e)}"}

    # ── Leitura ───────────────────────────────────────────────────────

    def _read(self, action, filepath, sheet_name, cell, column,
              row, skip_header, var_name, openpyxl) -> dict:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = self._get_sheet(wb, sheet_name)
        if isinstance(ws, dict):  # erro
            return ws

        from blocks.browser.extract_text import ExtractTextBlock
        context = ExtractTextBlock._context

        if action == "read_cell":
            if not cell:
                return {"success": False, "message": "cell é obrigatório para read_cell. Ex: A1"}
            val = ws[cell].value
            context[var_name] = str(val) if val is not None else ""
            return {"success": True, "message": f"Lido {cell} → '{var_name}': {val}"}

        if action == "read_column":
            if not column:
                return {"success": False, "message": "column é obrigatório para read_column. Ex: A ou 1"}
            col_idx = self._col_to_idx(column)
            values  = []
            for i, row_cells in enumerate(ws.iter_rows()):
                if i == 0 and skip_header:
                    continue
                cell_val = row_cells[col_idx - 1].value if col_idx <= len(row_cells) else None
                if cell_val is not None:
                    values.append(str(cell_val))
            context[var_name]              = values
            context[f"{var_name}_total"]   = str(len(values))
            context[f"{var_name}_primeiro"] = values[0] if values else ""
            return {"success": True, "message": f"Coluna {column} → {len(values)} valor(es) em '{var_name}'"}

        if action == "read_row":
            if not row:
                return {"success": False, "message": "row é obrigatório para read_row. Ex: 1"}
            try:
                row_idx = int(row)
            except ValueError:
                return {"success": False, "message": f"row deve ser um número. Recebido: '{row}'"}
            values = [str(c.value) for c in ws[row_idx] if c.value is not None]
            context[var_name]            = values
            context[f"{var_name}_total"] = str(len(values))
            return {"success": True, "message": f"Linha {row} → {len(values)} célula(s) em '{var_name}'"}

        if action == "read_sheet":
            start = 2 if skip_header else 1
            rows  = []
            headers = [c.value for c in ws[1]] if skip_header else None
            for row_cells in ws.iter_rows(min_row=start, values_only=True):
                if headers:
                    rows.append(dict(zip(headers, row_cells)))
                else:
                    rows.append(list(row_cells))
            context[var_name]            = rows
            context[f"{var_name}_total"] = str(len(rows))
            return {"success": True, "message": f"Planilha lida → {len(rows)} linha(s) em '{var_name}'"}

    # ── Escrita ───────────────────────────────────────────────────────

    def _write(self, action, filepath, sheet_name, cell, column, value, openpyxl) -> dict:
        wb = openpyxl.load_workbook(filepath)
        ws = self._get_sheet(wb, sheet_name)
        if isinstance(ws, dict):
            return ws

        if action == "write_cell":
            if not cell:
                return {"success": False, "message": "cell é obrigatório para write_cell. Ex: B3"}
            ws[cell] = value
            wb.save(filepath)
            return {"success": True, "message": f"Escrito em {cell}: \"{str(value)[:60]}\""}

        if action == "append_row":
            if not value:
                return {"success": False, "message": "value é obrigatório para append_row. Separe valores com ;"}
            values = [v.strip() for v in str(value).split(";")]
            ws.append(values)
            wb.save(filepath)
            return {"success": True, "message": f"Linha adicionada com {len(values)} coluna(s): {values[:3]}"}

    def _create(self, filepath: str, openpyxl) -> dict:
        if not filepath:
            return {"success": False, "message": "filepath é obrigatório para create."}
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        wb = openpyxl.Workbook()
        wb.save(filepath)
        return {"success": True, "message": f"Arquivo criado: {filepath}"}

    # ── Helpers ───────────────────────────────────────────────────────

    def _get_sheet(self, wb, sheet_name: str):
        if not sheet_name:
            return wb.active
        if sheet_name not in wb.sheetnames:
            return {"success": False, "message": f"Planilha '{sheet_name}' não encontrada. Disponíveis: {wb.sheetnames}"}
        return wb[sheet_name]

    def _col_to_idx(self, column: str) -> int:
        """Converte 'A' → 1, 'B' → 2, '3' → 3."""
        if column.isdigit():
            return int(column)
        result = 0
        for char in column.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result
